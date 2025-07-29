import json
from openpyxl import load_workbook
from collections import OrderedDict
import pandas as pd

def aptis_reading_to_json(file_path):
    """
    Chuyển đổi file Excel APTIS Reading thành cấu trúc JSON
    
    Args:
        file_path (str): Đường dẫn đến file Excel
        
    Returns:
        dict: Dữ liệu đã được chuyển đổi sang JSON
    """
    # Load workbook
    wb = load_workbook(filename=file_path)
    
    # Khởi tạo cấu trúc dữ liệu
    data = {
        "part1": [],
        "part2": [],
        "part3": [],
        "part4": []
    }
    
    # Xử lý Part 1 (giữ nguyên)
    ws = wb["part1"]
    current_group = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            current_group = {
                "group": row[0],
                "questions": []
            }
            data["part1"].append(current_group)
        if row[1] is not None:
            question = {
                "sentence": row[1],
                "correct_answer": row[2],
                "options": [row[3], row[4], row[5]],
                "explain": row[6]
            }
            current_group["questions"].append(question)
    
    # Xử lý Part 2 (giữ nguyên)
    ws = wb["part2"]
    current_topic = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            current_topic = {
                "topic": row[0],
                "sentences": [],
                "explain": row[4]
            }
            data["part2"].append(current_topic)
        
        if row[1] is not None:
            sentence = {
                "key": row[1],
                "text": row[2],
                "is_example_first": row[3]
            }
            current_topic["sentences"].append(sentence)
    
    # Xử lý Part 3 (cấu trúc mới)
    ws = wb["part3"]
    current_topic = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Xử lý khi gặp topic mới
        if row[0]:
            # Lưu topic trước đó nếu có
            if current_topic is not None:
                data["part3"].append(current_topic)
            
            # Tạo topic mới với cấu trúc mới
            current_topic = {
                "topic": row[0],
                "questions": [],
                "person_A": row[3] if row[3] else "",
                "person_B": row[4] if row[4] else "",
                "person_C": row[5] if row[5] else "",
                "person_D": row[6] if row[6] else ""
            }
        
        # Thêm câu hỏi cho topic hiện tại
        if row[1]:
            question = {
                "text": row[1],
                "correct_answer": row[2],
                "explain": row[7]
            }
            current_topic["questions"].append(question)
    
    # Lưu topic cuối cùng
    if current_topic is not None:
        data["part3"].append(current_topic)
    
    # Xử lý Part 4 (cấu trúc mới)
    ws = wb["part4"]
    current_topic = None
    
    # Lấy tất cả các header từ cột D (bỏ qua hàng đầu tiên)
    headers = []
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        if row[0] and row[0].strip():
            headers.append(row[0])
            
    explains = []
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
        if row[0] and row[0].strip():
            explains.append(row[0])
    
    # Xử lý nội dung
    current_topic = None
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=3, values_only=True)):
        if row[0]:
            # Lưu topic trước đó nếu có
            if current_topic is not None:
                data["part4"].append(current_topic)
            
            # Tạo topic mới với cấu trúc mới
            current_topic = {
                "topic": row[0],
                "options": headers,
                "questions": []
            }
        
        # Thêm câu hỏi nếu có nội dung đoạn văn
        if row[1]:
            # Chuyển đổi correct_answer thành số nguyên
            try:
                correct_id = int(float(row[2])) - 1 if row[2] else None
            except (TypeError, ValueError):
                correct_id = None
            question = {
                "text": row[1],
                "correct_answer": correct_id,
                "explain": explains[i]
                
            }
            current_topic["questions"].append(question)
    
    # Lưu topic cuối cùng
    if current_topic is not None:
        data["part4"].append(current_topic)
    
    return data

# Ví dụ sử dụng

def aptis_listening_to_json(file_path):
    wb = load_workbook(file_path)
    result = OrderedDict()
    
    # Xử lý part1
    if 'part1' in wb.sheetnames:
        ws = wb['part1']
        part1_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Kiểm tra Question có tồn tại
                item = {
                    "question": row[0],
                    "audio_link": row[1],
                    "correct_answer": row[2],
                    "options": [str(row[3]), str(row[4]), str(row[5])],
                    "transcript": row[6],
                    "explain": row[7]
                }
                part1_data.append(item)
        result['part1'] = part1_data
    
    # Xử lý part2
    if 'part2' in wb.sheetnames:
        ws = wb['part2']
        part2_data = []
        
        # Lấy thông tin từ hàng đầu tiên có dữ liệu
        row1 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        
        # Lấy các options từ cột D đến I
        options = [str(opt) for opt in row1[3:9] if opt is not None]
        
        # Tạo object cho part2
        part2_obj = {
            "topic": row1[0],
            "audio_link": row1[1],
            "a": int(row1[2]) if isinstance(row1[2], float) else row1[2],
            "options": options,
            "transcript": row1[9],
            "explain": row1[10]
        }
        
        # Thêm các correct_answer từ các hàng tiếp theo
        letters = ['b', 'c', 'd']
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
            if i >= len(letters):
                break
            if row[2] is not None:
                # Chuyển đổi thành số nguyên nếu là float
                value = row[2]
                if isinstance(value, float):
                    value = int(value)
                part2_obj[letters[i]] = value
        part2_data.append(part2_obj)
        result['part2'] = part2_data
    
    # Xử lý part3
    if 'part3' in wb.sheetnames:
        ws = wb['part3']
        part3_data = []
        current_group = None
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Nếu có topic hoặc audio mới, tạo group mới
            if row[0] or row[3]:
                # Lưu group hiện tại nếu có
                if current_group:
                    part3_data.append(current_group)
                
                # Tạo group mới
                current_group = {
                    "topic": row[0],
                    "questions": [],
                    "correct_answers": [],
                    "audio_link": row[3],
                    "transcript": row[4],
                    "explains": []
                }
            
            # Thêm câu hỏi và câu trả lời vào group hiện tại
            if row[1]:
                current_group["questions"].append(row[1])
                current_group["correct_answers"].append(row[2])
                current_group["explains"].append(row[5])
        
        # Lưu group cuối cùng
        if current_group:
            part3_data.append(current_group)
        
        result['part3'] = part3_data
    
    # Xử lý part4 theo định dạng mới
    if 'part4' in wb.sheetnames:
        ws = wb['part4']
        part4_data = []
        current_group = None
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Nếu có topic hoặc audio mới, tạo group mới
            if row[0] or row[1]:
                # Lưu group hiện tại nếu có
                if current_group:
                    part4_data.append(current_group)
                
                # Tạo group mới
                current_group = {
                    "topic": row[0],
                    "audio_link": row[1],
                    "questions": [],
                    "correct_answers": [],
                    "options": [],
                    "transcript": row[7],
                    "explains": []
                }
            
            # Thêm câu hỏi, đáp án và lựa chọn vào group hiện tại
            if row[2]:
                current_group["questions"].append(row[2])
                current_group["correct_answers"].append(row[3])
                current_group["explains"].append(row[8])
                current_group["options"].append([
                    str(row[4]),
                    str(row[5]),
                    str(row[6])
                ])
        
        # Lưu group cuối cùng
        if current_group:
            part4_data.append(current_group)
        
        result['part4'] = part4_data
    
    return result


def aptis_speaking_to_json(file_path: str) -> str:
    """
    Đọc file Excel chứa đề thi Speaking và chuyển đổi thành chuỗi JSON
    theo cấu trúc yêu cầu: một object cho mỗi part.

    Args:
        file_path (str): Đường dẫn tới file Excel (.xlsx).

    Returns:
        str: Một chuỗi JSON có cấu trúc đại diện cho đề thi.
             Trả về chuỗi JSON rỗng '[]' nếu có lỗi hoặc file trống.
    """
    try:
        df = pd.read_excel(file_path)

        # Bước 1: Tiền xử lý dữ liệu
        grouping_cols = ['Topic', 'part', 'instruction']
        for col in grouping_cols:
            if col not in df.columns:
                raise ValueError(f"File Excel thiếu cột bắt buộc: '{col}'")
        
        df[grouping_cols] = df[grouping_cols].ffill()
        df['part'] = df['part'].astype(int)
        
        # Thay thế các giá trị NaN (ô trống) bằng None để JSON hiển thị là "null"
        # Điều này đặc biệt quan trọng cho các cột image_url
        df = df.where(pd.notna(df), None)

        image_cols = [col for col in df.columns if col.startswith('image_url_')]
        
        # Bước 2: Xây dựng cấu trúc JSON theo yêu cầu
        exam_parts = []
        for part_id, group in df.groupby('part'):
            # Lấy thông tin chung từ dòng đầu tiên của nhóm
            first_row = group.iloc[0]

            # Gom tất cả các câu hỏi trong nhóm thành một danh sách
            questions_list = group['question'].dropna().tolist()

            part_data = {
                "part": int(part_id),
                "topic": first_row['Topic'],
                "instruction": first_row['instruction'],
                "question": questions_list
            }
            
            # Thêm các trường image_url vào đối tượng
            # Logic: Lấy giá trị đầu tiên không rỗng cho mỗi cột image_url trong group
            # Thường thì chỉ có 1 giá trị trong cả nhóm
            for img_col in image_cols:
                # Tìm giá trị không phải None đầu tiên trong cột ảnh của nhóm này
                first_valid_image_url = group[img_col].dropna().unique()
                if len(first_valid_image_url) > 0:
                    part_data[img_col] = first_valid_image_url[0]
                else:
                    part_data[img_col] = None

            exam_parts.append(part_data)

        return json.dumps(exam_parts, indent=4, ensure_ascii=False)

    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file tại '{file_path}'")
        return "[]"
    except ValueError as ve:
        print(f"Lỗi dữ liệu: {ve}")
        return "[]"
    except Exception as e:
        print(f"Đã xảy ra lỗi không xác định: {e}")
        return "[]"

def aptis_writing_to_json(file_path: str) -> str:
    """
    Đọc file Excel chứa đề thi Writing và chuyển đổi thành chuỗi JSON.

    File Excel phải có các cột: 'Topic', 'part', 'Instruction', 'question'.

    Args:
        file_path (str): Đường dẫn tới file Excel (.xlsx).

    Returns:
        str: Một chuỗi JSON có cấu trúc đại diện cho đề thi.
             Trả về chuỗi JSON rỗng '[]' nếu có lỗi hoặc file trống.
    """
    try:
        # Đọc file Excel
        df = pd.read_excel(file_path)

        # --- Bước 1: Tiền xử lý dữ liệu ---
        # Lấp đầy các giá trị bị thiếu trong các cột nhóm (do merged cells)
        grouping_cols = ['Topic', 'part']
        df[grouping_cols] = df[grouping_cols].ffill()
        
        # Đảm bảo cột 'part' là kiểu số nguyên
        df['part'] = df['part'].astype(int)

        # Thay thế các ô trống (NaN) bằng None để JSON hiển thị là "null"
        df = df.where(pd.notna(df), None)

        # --- Bước 2: Xây dựng cấu trúc JSON ---
        exam_parts = []
        # Nhóm DataFrame theo 'part' để xử lý từng phần thi
        for part_id, group in df.groupby('part'):
            # Lấy thông tin chung từ dòng đầu tiên của nhóm
            first_row = group.iloc[0]
            
            # Gom tất cả các câu hỏi trong nhóm thành một danh sách
            questions_list = group['question'].dropna().tolist()

            part_data = {
                "part_id": int(part_id),
                "topic": first_row['Topic'],
                # Lấy instruction, nếu không có sẽ là None
                "instruction": first_row.get('Instruction'),
                "questions": questions_list
            }
            
            exam_parts.append(part_data)

        # Chuyển đổi list Python thành chuỗi JSON
        return exam_parts
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file tại '{file_path}'")
        return []
    except Exception as e:
        print(f"Đã xảy ra lỗi không xác định: {e}")
        return []



def aptis_g_v_to_json(file_path):
    wb = load_workbook(file_path)
    
    # Process part1
    ws_part1 = wb['part1']
    part1_data = []
    for row in ws_part1.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        answer_map = {'A': row[2], 'B': row[3], 'C': row[4]}
        correct_content = answer_map.get(row[1], '')
        item = {
            "question": row[0],
            "A": row[2],
            "B": row[3],
            "C": row[4],
            "correct_answer": correct_content
        }
        part1_data.append(item)
    
    # Process part2
    ws_part2 = wb['part2']
    part2_data = []
    current_group = None
    current_options = None
    
    # Lấy danh sách header A → I từ E đến M (index 4–12)
    option_headers = [cell.value for cell in ws_part2[1][4:14]]
    
    for row_idx in range(2, ws_part2.max_row + 1):
        row = [cell.value for cell in ws_part2[row_idx]]
        if not any(row):
            continue
        
        group_val = row[0]
        topic_val = row[1]
        question_val = row[2]
        correct_key = row[3]  # Chữ cái như A, B, ...
        
        # Nếu là dòng bắt đầu group mới
        if group_val:
            if current_group:
                part2_data.append(current_group)
            
            options = {}
            for col_idx, header in enumerate(option_headers, start=4):
                if col_idx < len(row):
                    options[header] = row[col_idx] or ""
            
            current_options = options
            current_group = {
                "group": group_val,
                "topic": topic_val,
                "options": current_options,
                "questions": []
            }
        
        # Thêm câu hỏi vào group hiện tại
        if current_group and question_val:
            correct_answer_text = current_options.get(correct_key, "")
            current_group["questions"].append({
                "question": question_val,
                "correct_answer": correct_answer_text
            })
    
    # Thêm group cuối cùng nếu còn
    if current_group:
        part2_data.append(current_group)
    
    return {"part1": part1_data, "part2": part2_data}


if __name__ == "__main__":
    json_data = aptis_g_v_to_json("C:/Users/admin/Downloads/G&V TEST 1.xlsx")
    print(json_data)
    
    # Lưu ra file JSON
    with open("C:/Users/admin/Desktop/aptis_web/aptis_speaking.json", "w", encoding="utf-8") as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
        
